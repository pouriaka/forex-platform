import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment 
from openpyxl.utils import get_column_letter 
import yfinance as yf
import pandas as pd
import requests



def read_cell_value(file_path, sheet_name, cell_address):
    # Load the workbook and the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Get the cell value
    cell_value = sheet[cell_address].value
    return cell_value


class Growth_rate:
    def __init__(self):
        self.file_path = 'D:\\text\\intrinsic value\\us_stock_data.xlsx'   # Path to your Excel file of stock data
        self.revenue_sheet = 'Revenue'
        self.gross_profit_sheet = 'Gross_profit'
        self.operating_income = 'Operating_income'
        self.EBITDA = 'EBITDA' 
        self.Net_income_sheet = 'Net_income' 
        self.EPS_sheet = 'EPS'   
        self.dividen_sheet = 'Dividen'   
        self.FCF_sheet = 'FCF' 
        

    def net_income_growth_avg(self, symbol, from_year=2014, to_year=2023):
        # Read the Excel file
        df = pd.read_excel(self.file_path, sheet_name=self.Net_income_sheet)

        # Set 'Unnamed: 0' as the index and rename it to 'Year'
        df = df.set_index('Unnamed: 0')
        df.index.name = 'Year'

        # Filter data between tow years
        df = df.loc[from_year:to_year]

        # Calculate annual growth rates
        df['Growth'] = df[symbol].pct_change()

        # Calculate average annual net income growth rate(not pecent, convert to percent by excel)
        avg = df['Growth'].mean() 

        return avg


    def eps_growth_avg(self, symbol, from_year=2014, to_year=2023):
        # Read the Excel file
        df = pd.read_excel(self.file_path, sheet_name=self.EPS_sheet)

        # Set 'Unnamed: 0' as the index and rename it to 'Year'
        df = df.set_index('Unnamed: 0')
        df.index.name = 'Year'

        # Filter data between tow years
        df = df.loc[from_year:to_year]

        # Calculate annual growth rates
        df['Growth'] = df[symbol].pct_change()

        # Calculate average annual eps growth rate(not pecent, convert to percent by excel)
        avg = df['Growth'].mean() 

        return avg


    def dividen_growth_avg(self, symbol, from_year=2014, to_year=2023):
        # Read the Excel file
        df = pd.read_excel(self.file_path, sheet_name=self.dividen_sheet)

        # Set 'Unnamed: 0' as the index and rename it to 'Year'
        df = df.set_index('Unnamed: 0')
        df.index.name = 'Year'

        # Filter data between tow years
        df = df.loc[from_year:to_year]

        # Calculate annual growth rates
        df['Growth'] = df[symbol].pct_change()

        # Calculate average annual dividen growth rate(not pecent, convert to percent by excel)
        avg = df['Growth'].mean() 

        return avg



class DDM:
    def __init__(self):
        pass


    def dividen_growth_avg(self, symbol, year=12):
        # Fetch the stock data
        stock = yf.Ticker(symbol)

        # Get the dividend history
        dividends = stock.dividends

        # Convert dividends index to naive datetime (without timezone)
        dividends.index = dividends.index.tz_localize(None)

        # Filter dividends for the last 10 years
        start_date = pd.Timestamp.now() - pd.DateOffset(years=year)
        dividends_last_10_years = dividends[dividends.index >= start_date]

        # Resample the dividends to yearly frequency and sum them
        yearly_dividends = dividends_last_10_years.resample('Y').sum()
        yearly_dividends = yearly_dividends.drop(yearly_dividends.index[0])
        yearly_dividends = yearly_dividends.drop(yearly_dividends.index[-1])

        df = yearly_dividends

        # Calculate annual growth rates
        df['Dividend Growth'] = df.pct_change()

        # Calculate average annual dividend growth rate(not pecent, convert to percent by excel)
        average_growth_rate = df['Dividend Growth'].mean() 

        return average_growth_rate
    

    def current_dividen(self, symbol):
        # Fetch the stock data
        stock = yf.Ticker(symbol)

        # Get the dividend history
        dividends = stock.dividends

        # Convert dividends index to naive datetime (without timezone)
        dividends.index = dividends.index.tz_localize(None)

        # Filter dividends for the last 10 years
        start_date = pd.Timestamp.now() - pd.DateOffset(years=3)
        dividends_last_10_years = dividends[dividends.index >= start_date]

        # Resample the dividends to yearly frequency and sum them
        yearly_dividends = dividends_last_10_years.resample('Y').sum()
        yearly_dividends = yearly_dividends.drop(yearly_dividends.index[0])
        yearly_dividends = yearly_dividends.drop(yearly_dividends.index[-1])

        return yearly_dividends[-1]


    def beta(self, symbol):
        # Fetch the stock data
        stock = yf.Ticker(symbol)

        # Get the beta value
        beta = stock.info.get('beta')

        return beta


    # We use 10-year Treasury bond yield as risk free return
    def riskfree_return(self):
        # Define the ticker symbol for the 10-year Treasury bond yield
        treasury_symbol = '^TNX'

        # Fetch the data
        treasury = yf.Ticker(treasury_symbol)

        # Get the current yield
        treasury_yield = treasury.history(period='1d')['Close'].iloc[-1]
        treasury_yield = round(treasury_yield, 2) / 100
        

        return treasury_yield


    # We use S&P 500, Dow Jones return as market portfolio return
    def market_portfolio_return(self, year=10):
        # Define the ticker symbols for major indices representing the market portfolio
        index_symbols = ['^GSPC', '^DJI']  

        # Fetch historical data for the indices over the last 10 years
        start_date = pd.Timestamp.now() - pd.DateOffset(years=year)
        end_date = pd.Timestamp.now()
        indices_data = yf.download(index_symbols, start=start_date, end=end_date)['Adj Close']

        # Calculate annual returns for each index
        annual_returns = indices_data.resample('Y').ffill().pct_change().dropna()

        # Calculate equal-weighted market portfolio return (simple average of annual returns)
        market_portfolio_return = annual_returns.mean(axis=1).mean()  # Mean of mean annual returns

        # market portfolio is not percent(change to percent by excel)
        return market_portfolio_return


    def stock_price(self, symbol):
        # Fetch the stock data
        stock = yf.Ticker(symbol)

        # Get the latest stock price
        latest_price = stock.history(period='1d')['Close'][0]
        latest_price = round(latest_price, 2)
        
        return latest_price


    def excel_update(self, symbol_list):
        # Create a new Workbook
        wb = Workbook()
        ws = wb.create_sheet(title='DDM')
        # Specify the full path to save the workbook
        file_path = 'D:\\text\\intrinsic value\\DDM.xlsx'  

        # Name the column
        number = 1
        data_to_save = {
            f'A{number}': 'number',
            f'B{number}': 'stock_symbol',
            f'C{number}': 'stock_div_growth',
            f'D{number}': 'stock_beta',
            f'E{number}': 'risk_free_return',
            f'F{number}': 'market_return',
            f'G{number}': 'required_return',
            f'H{number}': 'ddm_price',
            f'I{number}': 'stock_price'
        }

        # Save data to specified cells
        for cell, value in data_to_save.items():
            ws[cell] = value
            # Center align the text in the merged cell
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            wb.save(file_path)


        # Take the stock information and calculate ddm price 
        for number in range(2, len(symbol_list) + 2):
            stock_symbol = symbol_list[number-2]
            stock_div_growth = self.dividen_growth_avg(stock_symbol)
            stock_beta = self.beta(stock_symbol)
            risk_free_return = self.riskfree_return()
            market_return = self.market_portfolio_return()
            required_return = risk_free_return + (stock_beta*(market_return - risk_free_return))
            stock_current_dividen = self.current_dividen(stock_symbol)
            D1 = stock_current_dividen + (stock_current_dividen * stock_div_growth)
            ddm_price =round(D1/(required_return - stock_div_growth), 2)
            stock_price = self.stock_price(stock_symbol)

            # Define data and its location
            data_to_save = {
                f'A{number}': number-1,
                f'B{number}': stock_symbol,
                f'C{number}': stock_div_growth,
                f'D{number}': stock_beta,
                f'E{number}': risk_free_return,
                f'F{number}': market_return,
                f'G{number}': required_return,
                f'H{number}': ddm_price,
                f'I{number}': stock_price
            }

            # Save data to specified cells
            for cell, value in data_to_save.items():
                ws[cell] = value
                # Center align the text in the merged cell
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
                wb.save(file_path)

                # Change the return cells to percent view
                # Define the cell addres to check
                characters_to_check = {'C', 'E', 'F', 'G'}
                cell_percent = ws[cell]
                # Check for the presence of any of the specified characters
                for char in characters_to_check:
                    if char == cell[0]:
                        cell_percent.number_format = '0.00%'  # Apply percentage format
                    
                wb.save(file_path)


        # Adjust all the cells wide to visible the data in it
        # Iterate over all columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get the column letter

            # Find the maximum length of any cell in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set the column width
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Get the sheet to be removed (It automaticly made when work sheet opened)
        sheet_to_remove = wb['Sheet']
        
        # Remove the sheet from the workbook
        wb.remove(sheet_to_remove)

        # Save the workbook
        wb.save(file_path)


        print("Data has been saved to DDM.xlsx in specified cells")



class Peter_Lynch:
    def __init__(self) -> None:
        self.file_path = 'D:\\text\\intrinsic value\\us_stock_data.xlsx'   # Path to your Excel file
        self.EPS_sheet = 'EPS'       # Name of the sheet you want to read from
        
    
    def stock_price(self, symbol):
        # Fetch the stock data
        stock = yf.Ticker(symbol)

        # Get the latest stock price
        latest_price = stock.history(period='1d')['Close'][0]
        latest_price = round(latest_price, 2)
        
        return latest_price


    def forward_dividend_yield(self, symbol):
        stock = yf.Ticker(symbol)
        info = stock.info
        
        forward_dividend_yield = info.get('dividendYield')
        
        if forward_dividend_yield is not None:
            return forward_dividend_yield * 100  # Convert to percentage
        else:
            return None


    def pe(self, symbol):
        # Fetch stock data using yfinance
        stock = yf.Ticker(symbol)
        
        # Get stock info
        stock_info = stock.info
        
        # Extract P/E ratio
        pe_ratio = stock_info.get('trailingPE')
        pe_ratio = round(pe_ratio, 2)
        
        return pe_ratio
    

    def eps_growth_avg(self, symbol, from_year=2014, to_year=2023):
        # Read the Excel file
        df = pd.read_excel(self.file_path, sheet_name=self.EPS_sheet)


        # Set 'Unnamed: 0' as the index and rename it to 'Year'
        df = df.set_index('Unnamed: 0')
        df.index.name = 'Year'

        # Filter data between tow years
        df = df.loc[from_year:to_year]

        # Calculate annual growth rates
        df['EPS Growth'] = df[symbol].pct_change()

        # Calculate average annual eps growth rate(not pecent, convert to percent by excel)
        avg_eps_growth_rate = df['EPS Growth'].mean() 


        return avg_eps_growth_rate


    def excel_update(self, symbol_list):
        # Create a new Workbook
        wb = Workbook()
        ws = wb.create_sheet(title='Peter_Lynch')
        # Specify the full path to save the workbook
        file_path = 'D:\\text\\intrinsic value\\Peter_Lynch.xlsx'   

        # Name the column
        number = 1
        data_to_save = {
            f'A{number}': 'number',
            f'B{number}': 'stock_symbol',
            f'C{number}': 'stock_eps_growth',
            f'D{number}': 'div_yield',
            f'E{number}': 'p/e',
            f'F{number}': 'peter_lynch_ratio',
            f'G{number}': 'peter_lynch_price',
            f'H{number}': 'stock_price'
        }

        # Save data to specified cells
        for cell, value in data_to_save.items():
            ws[cell] = value
            # Center align the text in the merged cell
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            wb.save(file_path)


        # Take the stock information and calculate ddm price 
        for number in range(2, len(symbol_list) + 2):
            stock_symbol = symbol_list[number-2]
            stock_eps_growth = round(self.eps_growth_avg(stock_symbol), 4)
            stock_div_yield = self.forward_dividend_yield(stock_symbol)
            stock_pe = self.pe(stock_symbol)
            stock_price = self.stock_price(stock_symbol)
            peter_lynch_ratio = round((((stock_eps_growth * 100) + stock_div_yield) / stock_pe), 2)
            peter_lynch_price = round((peter_lynch_ratio * stock_price), 2)

            # Define data and its location
            data_to_save = {
                f'A{number}': number-1,
                f'B{number}': stock_symbol,
                f'C{number}': stock_eps_growth,
                f'D{number}': stock_div_yield,
                f'E{number}': stock_pe,
                f'F{number}': peter_lynch_ratio,
                f'G{number}': peter_lynch_price,
                f'H{number}': stock_price
            }

            # Save data to specified cells
            for cell, value in data_to_save.items():
                ws[cell] = value
                # Center align the text in the merged cell
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
                wb.save(file_path)

                # Change the dividen yield cell to percent view
                cell_percent = ws[cell]
                
                if cell[0] == 'C':
                    cell_percent.number_format = '0.00%'  # Apply percentage format

                if cell[0] == 'D':
                    ws[cell] = value/100
                    cell_percent.number_format = '0.00%'  # Apply percentage format

                    
                    
                wb.save(file_path)


        # Adjust all the cells wide to visible the data in it
        # Iterate over all columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get the column letter

            # Find the maximum length of any cell in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set the column width
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width


        # Get the sheet to be removed (It automaticly made when work sheet opened)
        sheet_to_remove = wb['Sheet']
        
        # Remove the sheet from the workbook
        wb.remove(sheet_to_remove)

        # Save the workbook
        wb.save(file_path)

        print("Data has been saved to Peter_lynch.xlsx in specified cells")




class Benjamin_Graham:
    def __init__(self):
        self.fred_api_key = '4c95061fdbe964f5794b12691774c33b'


    @staticmethod
    def eps(symbol):
        # Fetch the stock data
        stock = yf.Ticker(symbol)
        
        # Get the key statistics
        stats = stock.info
        
        # Extract the EPS value
        eps = stats.get('trailingEps', 'EPS not found')
        
        return eps
    

    def AAA_corporate_bond_yield(self):
        # Correct series ID for Moody's Seasoned Aaa Corporate Bond Yield
        series_id = 'AAA'

        # Endpoint for the FRED API
        url = f'https://api.stlouisfed.org/fred/series/observations'

        # Parameters for the API request
        params = {
            'series_id': series_id,
            'api_key': self.fred_api_key,
            'file_type': 'json'
        }

        try:
            # Make a GET request to the API
            response = requests.get(url, params=params)
            response.raise_for_status()  # Raise an error for bad responses

            # Extract the most recent yield value
            data = response.json()
            if 'observations' in data:
                # The API returns data in reverse chronological order, so we take the first observation
                latest_yield = data['observations'][0]['value']

                return latest_yield
               
            else:
                print("No data available from the API.")

        except requests.exceptions.RequestException as e:
            print(f"Error fetching data: {e}")




stock_list = ['AAPL', 'MSFT','KO']

# DDM().excel_update(stock_list)
Peter_Lynch().excel_update(stock_list)


